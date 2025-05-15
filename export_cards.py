import openpyxl
import sys
import json

CARDS_FILE_PATH = 'cards.json'
SECRET_AGENDAS_FILE_PATH = 'secret_agendas.json'
TEMPLATE_PATH = "cards_template.xlsx"
TEMPLATE_ROWS = 4
OUTPUT_PATH = "cards.xlsx"


def load_json_data(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"Error: File not found - {file_path}", file=sys.stderr)
        return None
    except json.JSONDecodeError:
        print(f"Error: Could not decode JSON from - {file_path}", file=sys.stderr)
        return None


def read_cards(cards_path, secret_agendas_path):
    main_cards_data = load_json_data(cards_path)
    secret_agendas_data = load_json_data(secret_agendas_path)

    if main_cards_data is None or secret_agendas_data is None:
        print("Aborting due to file loading errors.", file=sys.stderr)
        return

    all_cards_for_csv = []

    for card in main_cards_data:
        title = card.get("title", "Без названия")
        description = card.get("description", "")
        count = card.get("count", 1)
        if not isinstance(count, int) or count < 1:
            count = 1  # Ensure count is at least 1

        for i in range(count):
            all_cards_for_csv.append([
                title,
                description,
            ])

    for agenda in secret_agendas_data:
        title = agenda.get("title", "Без названия")
        objective = agenda.get("objective_text", "")
        reward_text = agenda.get("reward_text", "")

        text = objective + " Тогда: \n" + reward_text

        all_cards_for_csv.append([
            title,
            text
        ])
    print("Read " + str(len(all_cards_for_csv)))
    return all_cards_for_csv


def write_cards(card_data, template_path, output_path):
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook.active

    print(f"Processing sheet: {sheet.title}")

    for card_index, (title, desc) in enumerate(card_data):
        card_row = (card_index % TEMPLATE_ROWS) * 2 + 1
        card_col = card_index // TEMPLATE_ROWS + 1
        cell_title = sheet.cell(card_row, card_col)
        cell_desc = sheet.cell(card_row + 1, card_col)
        cell_title.value = title
        cell_desc.value = desc

    workbook.save(output_path)
    workbook.close()
    print("Done writing " + output_path)


if __name__ == '__main__':
    cards = read_cards(CARDS_FILE_PATH, SECRET_AGENDAS_FILE_PATH)
    write_cards(cards, TEMPLATE_PATH, OUTPUT_PATH)
